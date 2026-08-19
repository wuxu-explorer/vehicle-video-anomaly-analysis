from __future__ import annotations

import os
import re
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config import (
    BASE_DIR, DATA_DIR, OUTPUT_DIR, REPORT_FILE,
    ALARM_TYPE_RULES, UNIT_RULES, INPUT_KEYWORDS,
    TREND_EPSILON, TOP_LIMITS,
)


def txt(v: Any) -> str:
    """安全地把单元格值转成文本；避免 pd.isna(list) 产生布尔歧义。"""
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except (TypeError, ValueError):
        pass
    return str(v).strip()


def norm(s: Any) -> str:
    return re.sub(r"\s+", "", txt(s)).lower()


def row_to_text(row: pd.Series) -> str:
    """把一整行安全拼接成文本。修复 sequence item expected str, float found。"""
    return " ".join(txt(v) for v in row.tolist() if txt(v))


def parse_date_from_text(s: Any):
    s = txt(s)
    patterns = [
        r"(20\d{2})[-_/年](\d{1,2})[-_/月](\d{1,2})",
        r"(20\d{2})[-_/年](\d{1,2})月",
        r"(20\d{2})(\d{2})(\d{2})",
    ]
    for p in patterns:
        m = re.search(p, s)
        if m:
            try:
                y, mo = int(m.group(1)), int(m.group(2))
                d = int(m.group(3)) if len(m.groups()) >= 3 else 1
                return pd.Timestamp(y, mo, d)
            except Exception:
                continue
    return pd.NaT


def detect_period_from_workbook(path: Path) -> pd.Timestamp:
    """优先读取Excel内部日期/时间，再看文件名；不依赖文件修改时间。"""
    try:
        xls = pd.ExcelFile(path)
        best = []
        for sheet in xls.sheet_names:
            try:
                raw = pd.read_excel(path, sheet_name=sheet, nrows=200)
            except Exception:
                continue
            if raw.empty:
                continue

            # 先检查明显的日期/时间列。
            candidate_cols = []
            for col in raw.columns:
                cname = norm(col)
                if any(k in cname for k in ["日期", "时间", "开始时间", "结束时间", "报警时间", "事件时间"]):
                    candidate_cols.append(col)

            for col in candidate_cols:
                vals = pd.to_datetime(raw[col], errors="coerce")
                vals = vals[vals.notna()]
                if not vals.empty:
                    best.append(vals.min().normalize())

            # 再从单元格文本中找 2026-07 / 2026年7月 等日期。
            values = raw.astype(object).to_numpy().ravel().tolist()[:2000]
            for value in values:
                d = parse_date_from_text(value)
                if pd.notna(d):
                    best.append(d)

        if best:
            return min(best)
    except Exception:
        pass

    # Excel内部完全没有日期时，文件名只作为第二层兜底。
    d = parse_date_from_text(path.stem)
    if pd.notna(d):
        return d
    m = re.search(r"(20\d{2})[-_/年](\d{1,2})(?:月)?", path.stem)
    if m:
        try:
            return pd.Timestamp(int(m.group(1)), int(m.group(2)), 1)
        except Exception:
            pass
    return pd.NaT


def find_col(df: pd.DataFrame, aliases: list[str]) -> str | None:
    cmap = {norm(c): c for c in df.columns}
    for a in aliases:
        if norm(a) in cmap:
            return cmap[norm(a)]
    for c in df.columns:
        nc = norm(c)
        if any(norm(a) in nc for a in aliases):
            return c
    return None


def load_workbook_data(path: Path):
    """自动从一个Excel中识别里程表和报警明细表。"""
    xls = pd.ExcelFile(path)
    mileage = None
    alarms = None
    all_sheets: list[tuple[str, pd.DataFrame]] = []

    for sheet in xls.sheet_names:
        try:
            df = pd.read_excel(path, sheet_name=sheet)
        except Exception:
            continue
        if df.empty:
            continue
        all_sheets.append((sheet, df))

        mileage_col = find_col(df, ["行驶里程(km)", "行驶里程", "运行里程", "里程(km)", "总里程"])
        plate_col = find_col(df, ["车牌号码", "车牌", "车辆编号", "设备编号"])
        group_col = find_col(df, ["归属分组", "归属车组", "车组", "班组"])
        if mileage_col and plate_col:
            temp = df.copy()
            temp = temp.rename(columns={mileage_col: "_mileage", plate_col: "_plate"})
            temp["_group"] = temp[group_col].map(txt) if group_col else ""
            temp["_mileage"] = pd.to_numeric(
                temp["_mileage"].astype(str).str.replace(",", "", regex=False),
                errors="coerce",
            ).fillna(0)
            temp["_plate"] = temp["_plate"].map(txt)
            mileage = temp[["_plate", "_group", "_mileage"]].copy()

        state_col = find_col(df, ["状态名称", "报警名称", "报警类型", "事件类型", "事件名称"])
        time_col = find_col(df, ["状态开始时间", "报警时间", "事件时间", "开始时间", "时间"])
        content_col = find_col(df, ["状态内容", "报警内容", "事件内容", "内容"])
        if state_col and (time_col or content_col or find_col(df, ["状态类型", "类型"])):
            temp = df.copy().rename(columns={state_col: "_alarm"})
            plate_col2 = find_col(temp, ["车牌号码", "车牌", "车辆编号", "设备编号"])
            group_col2 = find_col(temp, ["归属分组", "归属车组", "车组", "班组"])
            temp["_plate"] = temp[plate_col2].map(txt) if plate_col2 else ""
            temp["_group"] = temp[group_col2].map(txt) if group_col2 else ""
            temp["_alarm"] = temp["_alarm"].map(txt)
            # 关键修复：不能使用 .agg(" ".join, axis=1)，因为行中可能同时存在float/str。
            temp["_alltext"] = temp.apply(row_to_text, axis=1)
            temp["_time"] = temp[time_col] if time_col else pd.NaT
            alarms = temp

    if mileage is None:
        for _, df in all_sheets:
            mc = find_col(df, ["行驶里程(km)", "行驶里程", "运行里程", "里程(km)", "总里程"])
            pc = find_col(df, ["车牌号码", "车牌", "车辆编号", "设备编号"])
            if mc and pc:
                gc = find_col(df, ["归属分组", "归属车组", "车组", "班组"])
                mileage = pd.DataFrame({
                    "_plate": df[pc].map(txt),
                    "_group": df[gc].map(txt) if gc else "",
                    "_mileage": pd.to_numeric(df[mc], errors="coerce").fillna(0),
                })
                break

    if alarms is None:
        for _, df in all_sheets:
            ac = find_col(df, ["报警名称", "报警类型", "事件类型", "事件名称", "状态名称"])
            if ac:
                alarms = df.copy().rename(columns={ac: "_alarm"})
                pc = find_col(df, ["车牌号码", "车牌", "车辆编号", "设备编号"])
                gc = find_col(df, ["归属分组", "归属车组", "车组", "班组"])
                alarms["_plate"] = df[pc].map(txt) if pc else ""
                alarms["_group"] = df[gc].map(txt) if gc else ""
                alarms["_alltext"] = df.apply(row_to_text, axis=1)
                tc = find_col(df, ["状态开始时间", "报警时间", "事件时间", "开始时间", "时间"])
                alarms["_time"] = df[tc] if tc else pd.NaT
                break

    return mileage, alarms


def classify_alarm(row: pd.Series) -> str | None:
    s = norm(" ".join([txt(row.get("_alarm", "")), txt(row.get("_alltext", ""))]))
    for name, keywords in ALARM_TYPE_RULES:
        if any(norm(k) in s for k in keywords):
            return name
    return None


def unit_of(group: Any) -> str:
    g = norm(group)
    for unit, keywords in UNIT_RULES.items():
        if any(norm(k) in g for k in keywords):
            return unit
    return "其他"


def analyze_period(path: Path):
    mileage, alarms = load_workbook_data(path)
    period = detect_period_from_workbook(path)

    if mileage is None:
        mileage = pd.DataFrame(columns=["_plate", "_group", "_mileage"])
    mileage = mileage.copy()
    mileage["_mileage"] = pd.to_numeric(mileage["_mileage"], errors="coerce").fillna(0)
    mileage = mileage[mileage["_mileage"] > 0].copy()
    mileage["_plate"] = mileage["_plate"].map(txt)
    mileage["_group"] = mileage["_group"].map(txt)

    total_vehicles = mileage.loc[mileage["_plate"] != "", "_plate"].nunique()
    total_mileage = float(mileage["_mileage"].sum())

    if alarms is None:
        alarms = pd.DataFrame(columns=["_alarm", "_plate", "_group", "_alltext", "_time"])
    alarms = alarms.copy()
    alarms["_alarm_type"] = alarms.apply(classify_alarm, axis=1)
    classified = alarms[alarms["_alarm_type"].notna()].copy()

    # 如果有报警明细，但本次关键词没有命中，则仍保留“其他报警”，避免总报警数变成0。
    total_alarm = len(classified)
    if total_alarm == 0 and len(alarms) > 0:
        total_alarm = len(alarms)
        alarms["_alarm_type"] = alarms["_alarm"].map(lambda x: txt(x) or "其他报警")
        classified = alarms.copy()

    classified["_unit"] = classified["_group"].map(unit_of)
    counts = classified["_alarm_type"].value_counts().to_dict()
    risk_per_100km = (total_alarm / total_mileage * 100) if total_mileage > 0 else np.nan

    unit_df = (
        classified.groupby("_unit", dropna=False).size().reset_index(name="报警次数")
        .sort_values("报警次数", ascending=False)
    )
    unit_df["占比"] = unit_df["报警次数"] / total_alarm if total_alarm else 0.0

    return {
        "path": path,
        "period": period,
        "vehicles": int(total_vehicles),
        "mileage": total_mileage,
        "alarm_total": int(total_alarm),
        "risk100": risk_per_100km,
        "alarm_counts": counts,
        "classified": classified,
        "unit_df": unit_df,
    }


def choose_two_files() -> tuple[Path, Path]:
    files = [p for p in DATA_DIR.glob("*.xlsx") if not p.name.startswith("~$")]
    if len(files) < 2:
        raise FileNotFoundError("data 文件夹至少需要放两份 Excel：旧数据和新数据。")

    infos = [(p, detect_period_from_workbook(p)) for p in files]
    dated = [(p, d) for p, d in infos if pd.notna(d)]
    if len(dated) >= 2:
        # 按Excel内部时间排序；同一时间时再按文件名排序，保证结果稳定。
        dated.sort(key=lambda x: (x[1], x[0].name.lower()))
        return dated[-2][0], dated[-1][0]

    # 只有在所有Excel都没有任何可识别内部日期、且文件名也无日期时才使用修改时间兜底。
    files.sort(key=lambda p: p.stat().st_mtime)
    return files[-2], files[-1]


def fmt_int(v):
    if pd.isna(v): return "—"
    return f"{int(round(float(v))):,}"


def fmt_km(v):
    if pd.isna(v): return "—"
    return f"{float(v):,.2f}"


def fmt_pct(v):
    if pd.isna(v): return "—"
    return f"{float(v) * 100:.2f}%"


def trend(new, old, epsilon=TREND_EPSILON):
    if pd.isna(new) or pd.isna(old): return "无法比较"
    diff = float(new) - float(old)
    base = max(abs(float(old)), 1e-12)
    if abs(diff) / base <= epsilon: return "基本持平"
    return "上升" if diff > 0 else "下降"


def diff_text(new, old, formatter=fmt_int):
    if pd.isna(new) or pd.isna(old): return "—"
    diff = float(new) - float(old)
    if diff == 0: return "0"
    sign = "+" if diff > 0 else "-"
    return sign + formatter(abs(diff))


def build_report(old, new):
    overview = pd.DataFrame([
        ["总运行车辆数", fmt_int(old["vehicles"]), fmt_int(new["vehicles"]), diff_text(new["vehicles"], old["vehicles"]), trend(new["vehicles"], old["vehicles"])],
        ["总运行里程（KM）", fmt_km(old["mileage"]), fmt_km(new["mileage"]), diff_text(new["mileage"], old["mileage"], fmt_km), trend(new["mileage"], old["mileage"])],
        ["总报警次数", fmt_int(old["alarm_total"]), fmt_int(new["alarm_total"]), diff_text(new["alarm_total"], old["alarm_total"]), trend(new["alarm_total"], old["alarm_total"])],
        ["百公里风险事件数", fmt_km(old["risk100"]), fmt_km(new["risk100"]), diff_text(new["risk100"], old["risk100"], fmt_km), trend(new["risk100"], old["risk100"])],
    ], columns=["指标", "旧数据", "新数据", "变化", "趋势"])

    rows = []
    for name in [r[0] for r in ALARM_TYPE_RULES]:
        oc = int(old["alarm_counts"].get(name, 0))
        nc = int(new["alarm_counts"].get(name, 0))
        op = oc / old["alarm_total"] if old["alarm_total"] else 0.0
        npct = nc / new["alarm_total"] if new["alarm_total"] else 0.0
        rows.append([name, oc, nc, nc - oc, fmt_pct(op), fmt_pct(npct), trend(npct, op)])
    alarm_trend = pd.DataFrame(rows, columns=["报警类型", "旧次数", "新次数", "次数变化", "旧占比", "新占比", "占比趋势"])

    # 单位取新旧并集，避免“旧数据存在但新数据没有”的单位直接消失。
    old_units = old["unit_df"].set_index("_unit")["报警次数"].to_dict() if not old["unit_df"].empty else {}
    new_units = new["unit_df"].set_index("_unit")["报警次数"].to_dict() if not new["unit_df"].empty else {}
    unit_rows = []
    for unit_name in sorted(set(old_units) | set(new_units), key=lambda u: (-new_units.get(u, 0), str(u))):
        oc = int(old_units.get(unit_name, 0)); nc = int(new_units.get(unit_name, 0))
        share = nc / new["alarm_total"] if new["alarm_total"] else 0.0
        unit_rows.append([unit_name, nc, fmt_pct(share), oc, nc - oc, trend(nc, oc)])
    unit = pd.DataFrame(unit_rows, columns=["单位", "新报警次数", "占总报警比例", "旧报警次数", "变化", "趋势"])

    c = new["classified"]
    top_tables = {}
    mapping = [
        ("车距过近高频违规车组TOP5", "车距过近", TOP_LIMITS.get("车距过近", 5)),
        ("疲劳驾驶高频违规车组TOP10", "疲劳驾驶", TOP_LIMITS.get("疲劳驾驶", 10)),
        ("打电话和抽烟违规车组TOP5", ["违规打电话", "违规抽烟"], TOP_LIMITS.get("打电话和抽烟", 5)),
        ("前车碰撞高频违规车组TOP5", "碰撞报警", TOP_LIMITS.get("碰撞报警", 5)),
        ("左右盲区违规车组TOP5", "左右盲区报警", TOP_LIMITS.get("左右盲区报警", 5)),
        ("驾驶员分心驾驶高频违规车组TOP10", "驾驶员分心", TOP_LIMITS.get("驾驶员分心", 10)),
    ]
    for title, types, limit in mapping:
        q = c[c["_alarm_type"].isin(types)] if isinstance(types, list) else c[c["_alarm_type"] == types]
        t = (q.groupby("_group", dropna=False).size().reset_index(name="预警次数")
             .sort_values(["预警次数", "_group"], ascending=[False, True], na_position="last").head(limit))
        t.insert(0, "排名", range(1, len(t) + 1))
        t = t.rename(columns={"_group": "车组"})
        top_tables[title] = t
    return overview, alarm_trend, unit, top_tables


def style_sheet(ws, header_rows: set[int] | None = None):
    header_rows = header_rows or {1}
    thin = Side(style="thin", color="D9E1F2")
    for r in header_rows:
        for cell in ws[r]:
            if cell.value not in (None, ""):
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")
    for i, col in enumerate(ws.columns, 1):
        width = max(10, min(35, max(len(str(c.value or "")) for c in col) + 2))
        ws.column_dimensions[get_column_letter(i)].width = width


def write_report(old, new):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    target = OUTPUT_DIR / REPORT_FILE
    temp = OUTPUT_DIR / f".__tmp_{REPORT_FILE}"
    if temp.exists():
        temp.unlink()

    overview, alarm_trend, unit, top_tables = build_report(old, new)

    try:
        with pd.ExcelWriter(temp, engine="openpyxl") as writer:
            overview.to_excel(writer, sheet_name="一_车辆运营里程", index=False, startrow=3)
            ws = writer.book["一_车辆运营里程"]
            ws["A1"] = "第一部分 车辆运营里程"
            ws["A1"].font = Font(size=16, bold=True)
            ws["A2"] = f"旧数据：{old['path'].name}    新数据：{new['path'].name}"
            ws["A2"].font = Font(italic=True, color="666666")

            alarm_trend.to_excel(writer, sheet_name="二_报警趋势分析", index=False, startrow=3)
            ws2 = writer.book["二_报警趋势分析"]
            ws2["A1"] = "第二部分 报警趋势分析"
            ws2["A1"].font = Font(size=16, bold=True)
            ws2["A2"] = "占比均以百分数显示；趋势为新数据相对旧数据的变化。"
            ws2["A2"].font = Font(italic=True, color="666666")
            start = 5 + len(alarm_trend) + 2
            ws2.cell(start, 1, "各单位报警数据").font = Font(size=13, bold=True)
            unit.to_excel(writer, sheet_name="二_报警趋势分析", index=False, startrow=start)

            ws3 = writer.book.create_sheet("三_典型违规案例")
            ws3["A1"] = "第三部分 典型违规案例晾晒"
            ws3["A1"].font = Font(size=16, bold=True)
            row = 3
            for title, table in top_tables.items():
                ws3.cell(row, 1, title).font = Font(size=12, bold=True)
                row += 1
                table.to_excel(writer, sheet_name="三_典型违规案例", index=False, startrow=row - 1)
                row += len(table) + 3

            # 明确保证至少有一个可见工作表。
            for wsx in writer.book.worksheets:
                wsx.sheet_state = "visible"
            writer.book.active = 0

        # 完整关闭后再打开做格式化，避免“至少一个sheet可见”等保存异常。
        wb = load_workbook(temp)
        for wsx in wb.worksheets:
            if wsx.title == "一_车辆运营里程":
                style_sheet(wsx, {4})
            elif wsx.title == "二_报警趋势分析":
                style_sheet(wsx, {4, 5 + len(alarm_trend) + 2})
            else:
                header_rows = set()
                r = 4
                for table in top_tables.values():
                    header_rows.add(r)
                    r += len(table) + 4
                style_sheet(wsx, header_rows)
            wsx.sheet_state = "visible"
        wb.active = 0
        wb.save(temp)

        # 原文件被Excel占用时，不覆盖失败；自动另存为新文件。
        try:
            os.replace(temp, target)
            final = target
        except PermissionError:
            alt = OUTPUT_DIR / f"{target.stem}_新生成{target.suffix}"
            os.replace(temp, alt)
            final = alt
        return final
    except Exception:
        if temp.exists():
            try: temp.unlink()
            except Exception: pass
        raise


def main():
    print("=" * 70)
    print("车辆运营与报警月度报告 V8 精简版（修正版）")
    print("自动读取Excel内部时间，比较最新两份数据")
    print("=" * 70)
    old_file, new_file = choose_two_files()
    print(f"旧数据：{old_file.name}")
    print(f"新数据：{new_file.name}")
    old = analyze_period(old_file)
    new = analyze_period(new_file)
    print(f"旧数据：车辆 {old['vehicles']}，里程 {old['mileage']:.2f} KM，报警 {old['alarm_total']}")
    print(f"新数据：车辆 {new['vehicles']}，里程 {new['mileage']:.2f} KM，报警 {new['alarm_total']}")
    out = write_report(old, new)
    print(f"结果：{out}")


if __name__ == "__main__":
    main()

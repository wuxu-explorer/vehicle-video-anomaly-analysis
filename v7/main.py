# -*- coding: utf-8 -*-

from __future__ import annotations

import os
import re
import time
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd

from config import (
    INPUT_FILE,
    OUTPUT_DIR,
    AUTO_DISCOVER_INPUT,
    COLUMN_ALIASES,
    REQUIRED_COLUMNS,
    CAMERA_ERROR_NAMES,
    GROUP_CHANNEL_RULES,
    DEFAULT_CHANNEL_RULES,
    CORE_SYSTEMS,
    SYSTEM_LAYER,
    SYSTEM_WEIGHT,
    REVIEW_SYSTEMS,
    RISK_WEIGHTS,
    PRIORITY_SHARE,
    PRIORITY_ORDER,
    PRIORITY_ADVICE,
    TOP_FOCUS_COUNT,
    TOP100_COUNT,
    OUTPUT_FILE_NAME,
    TEMP_FILE_NAME,
)


# ============================================================
# 基础工具
# ============================================================

def text(value: Any) -> str:

    if value is None:
        return ""

    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass

    return str(value).strip()


# ============================================================
# 通道号提取
# ============================================================

def channel_from_content(value: Any) -> int | None:

    s = text(value)

    if not s:
        return None

    patterns = [

        r"通道号\s*[:：#]?\s*(\d+)",

        r"通道\s*[:：#]?\s*(\d+)",

        r"channel\s*[:：#_\-]?\s*(\d+)",

        r"ch(?:annel)?\s*[:：#_\-]?\s*(\d+)",

        r"\bCH\s*[-_:#：]?\s*(\d+)\b",

        r"(\d+)\s*通道\b",
    ]

    for pattern in patterns:

        match = re.search(
            pattern,
            s,
            flags=re.IGNORECASE,
        )

        if match:

            return int(match.group(1))

    return None


# ============================================================
# 安全转换整数
# ============================================================

def safe_int(value: Any) -> int | None:

    s = text(value)

    if not s:
        return None

    try:

        return int(float(s))

    except (TypeError, ValueError):

        return channel_from_content(s)


# ============================================================
# 自动寻找输入文件
# ============================================================

def find_input_file() -> Path:

    if INPUT_FILE.exists():

        return INPUT_FILE

    if AUTO_DISCOVER_INPUT:

        candidates = [

            p
            for p in INPUT_FILE.parent.glob("*.xlsx")

            if not p.name.startswith("~$")

            and not p.name.startswith("视频异常巡检最终汇总")
        ]

        if candidates:

            return max(
                candidates,
                key=lambda p: p.stat().st_mtime,
            )

    raise FileNotFoundError(
        f"""
找不到输入Excel：

{INPUT_FILE}

请把原始巡检报表放入 data 文件夹。
"""
    )


# ============================================================
# Excel字段标准化
# ============================================================

def canonicalize_columns(
    df: pd.DataFrame,
) -> pd.DataFrame:

    df = df.copy()

    df.columns = [
        text(column)
        for column in df.columns
    ]

    lower_to_actual = {

        column.lower(): column

        for column in df.columns
    }

    rename_map = {}

    for canonical, aliases in COLUMN_ALIASES.items():

        if canonical in df.columns:

            continue

        for alias in aliases:

            actual = lower_to_actual.get(
                alias.lower()
            )

            if actual is not None:

                rename_map[actual] = canonical

                break

    if rename_map:

        df = df.rename(
            columns=rename_map
        )

    missing = [

        column

        for column in REQUIRED_COLUMNS

        if column not in df.columns
    ]

    if missing:

        raise ValueError(
            f"""
输入报表缺少必要字段：

{missing}

当前字段：

{list(df.columns)}
"""
        )

    for column in REQUIRED_COLUMNS:

        df[column] = df[column].map(text)

    if "通道号" in df.columns:

        df["通道号"] = (
            df["通道号"]
            .map(safe_int)
        )

    return df


# ============================================================
# 读取Excel
# ============================================================

def load_data():

    path = find_input_file()

    df = pd.read_excel(path)

    df = canonicalize_columns(df)

    return df, path


# ============================================================
# DMS / ADAS / DSC / 普通摄像头分类
# ============================================================

def classify(
    group: Any,
    channel: Any,
):

    group_name = text(group)

    channel_number = safe_int(channel)

    # --------------------------------------------------------
    # 无通道号
    # --------------------------------------------------------

    if channel_number is None:

        return (
            "未知",
            "无法提取通道号",
            "待确认",
        )

    # --------------------------------------------------------
    # 先匹配特殊车组
    # --------------------------------------------------------

    for rule in GROUP_CHANNEL_RULES:

        if rule["pattern"] in group_name:

            if channel_number in rule["channels"]:

                system = (
                    rule["channels"]
                    [channel_number]
                )

                return (
                    system,
                    f"车组专属规则：{rule['name']}",
                    "已确认",
                )

            return (
                "未配置",
                f"车组“{rule['name']}”未覆盖通道{channel_number}",
                "待确认",
            )

    # --------------------------------------------------------
    # 再使用通用通道规则
    # --------------------------------------------------------

    if channel_number in DEFAULT_CHANNEL_RULES:

        system = (
            DEFAULT_CHANNEL_RULES
            [channel_number]
        )

        return (
            system,
            "通用通道规则",
            "已确认",
        )

    # --------------------------------------------------------
    # 完全未知
    # --------------------------------------------------------

    return (
        "未知",
        f"规则库未覆盖通道{channel_number}",
        "待确认",
    )


# ============================================================
# 提取视频异常
# ============================================================

def extract_camera_errors(
    raw: pd.DataFrame,
):

    empty_columns = [

        "设备编号",

        "归属车组",

        "状态名称",

        "状态类型",

        "通道号",

        "系统类型",

        "系统层级",

        "分类依据",

        "规则状态",

        "状态内容",
    ]

    names = {
        text(name)
        for name in CAMERA_ERROR_NAMES
    }

    mask = raw["状态名称"].isin(names)

    mask = (
        mask
        |
        raw["状态名称"].str.contains(
            "摄像头|视频丢失|视频异常|视频丢帧|视频中断",
            case=False,
            na=False,
        )
    )

    mask = (
        mask
        |
        raw["状态内容"].str.contains(
            "摄像头异常|视频丢失|视频异常|视频丢帧|视频中断",
            case=False,
            na=False,
        )
    )

    camera = raw.loc[mask].copy()

    # --------------------------------------------------------
    # 没有视频异常
    # --------------------------------------------------------

    if camera.empty:

        return pd.DataFrame(
            columns=empty_columns
        )

    # --------------------------------------------------------
    # 获取通道号
    # --------------------------------------------------------

    if "通道号" in camera.columns:

        camera["通道号"] = (
            camera["通道号"]
            .map(safe_int)
        )

        extracted = (
            camera["状态内容"]
            .map(channel_from_content)
        )

        camera["通道号"] = (
            camera["通道号"]
            .where(
                camera["通道号"].notna(),
                extracted,
            )
        )

    else:

        camera["通道号"] = (
            camera["状态内容"]
            .map(channel_from_content)
        )

    # --------------------------------------------------------
    # 系统分类
    # --------------------------------------------------------

    classified = camera.apply(

        lambda row: classify(
            row["归属车组"],
            row["通道号"],
        ),

        axis=1,

        result_type="expand",
    )

    classified.columns = [

        "系统类型",

        "分类依据",

        "规则状态",
    ]

    camera[
        [
            "系统类型",
            "分类依据",
            "规则状态",
        ]
    ] = classified

    # --------------------------------------------------------
    # 系统层级
    # --------------------------------------------------------

    camera["系统层级"] = (

        camera["系统类型"]

        .map(SYSTEM_LAYER)

        .fillna("待确认")
    )

    return camera[
        empty_columns
    ].copy()


# ============================================================
# 统计车组 + 通道
# ============================================================

def build_stats(
    detail: pd.DataFrame,
):

    columns = [

        "归属车组",

        "通道号",

        "系统类型",

        "系统层级",

        "分类依据",

        "规则状态",

        "视频丢失次数",
    ]

    if detail.empty:

        return pd.DataFrame(
            columns=columns
        )

    stats = (

        detail

        .groupby(
            [
                "归属车组",

                "通道号",

                "系统类型",

                "系统层级",

                "分类依据",

                "规则状态",
            ],

            dropna=False,
        )

        .size()

        .reset_index(
            name="视频丢失次数"
        )
    )

    return stats


# ============================================================
# 归一化
# ============================================================

def normalize(
    series: pd.Series,
):

    series = pd.to_numeric(
        series,
        errors="coerce",
    ).fillna(0.0)

    if len(series) == 0:

        return series.astype(float)

    low = float(
        series.min()
    )

    high = float(
        series.max()
    )

    if high <= low:

        return pd.Series(
            0.5,
            index=series.index,
            dtype=float,
        )

    return (
        (series - low)
        /
        (high - low)
    ).clip(0, 1)


# ============================================================
# 百分位排名
# ============================================================

def pct_rank(
    series: pd.Series,
):

    if len(series) <= 1:

        return pd.Series(
            1.0,
            index=series.index,
            dtype=float,
        )

    return series.rank(
        method="average",
        pct=True,
    )


# ============================================================
# 动态数量
# ============================================================

def dynamic_cut(
    number: int,
    share: float,
):

    if number <= 0:

        return 0

    return max(
        1,
        int(
            np.ceil(
                number * share
            )
        ),
    )


# ============================================================
# 同一层级内部风险评分
# ============================================================

def score_within_layer(
    layer_df: pd.DataFrame,
):

    df = layer_df.copy()

    if df.empty:

        return df

    # --------------------------------------------------------
    # 每个车组总异常次数
    # --------------------------------------------------------

    group_total = (

        df

        .groupby(
            "归属车组",
            dropna=False,
        )["视频丢失次数"]

        .sum()

        .rename(
            "车组异常总次数"
        )
    )

    # --------------------------------------------------------
    # 每个系统总异常次数
    # --------------------------------------------------------

    system_total = (

        df

        .groupby(
            "系统类型",
            dropna=False,
        )["视频丢失次数"]

        .sum()

        .rename(
            "系统异常总次数"
        )
    )

    # --------------------------------------------------------
    # 每个车组异常通道数
    # --------------------------------------------------------

    spread = (

        df

        .groupby(
            "归属车组",
            dropna=False,
        )["通道号"]

        .nunique(
            dropna=True
        )

        .rename(
            "车组异常通道数"
        )
    )

    df = df.join(
        group_total,
        on="归属车组",
    )

    df = df.join(
        system_total,
        on="系统类型",
    )

    df = df.join(
        spread,
        on="归属车组",
    )

    # --------------------------------------------------------
    # 系统权重
    # --------------------------------------------------------

    df["系统权重"] = (

        df["系统类型"]

        .map(SYSTEM_WEIGHT)

        .fillna(0.0)
    )

    # --------------------------------------------------------
    # 各项风险分
    # --------------------------------------------------------

    df["次数排名分"] = pct_rank(
        df["视频丢失次数"]
    )

    df["绝对次数分"] = normalize(
        df["视频丢失次数"]
    )

    df["车组影响分"] = normalize(
        df["车组异常总次数"]
    )

    df["系统影响分"] = normalize(
        df["系统异常总次数"]
    )

    df["扩散分"] = normalize(
        df["车组异常通道数"]
    )

    df["系统权重分"] = (
        df["系统权重"]
    )

    # --------------------------------------------------------
    # 综合风险
    # --------------------------------------------------------

    weights = RISK_WEIGHTS

    total_weight = sum(
        weights.values()
    )

    df["动态风险分"] = (

        df["次数排名分"]
        * weights["count_rank"]

        +

        df["绝对次数分"]
        * weights["count_absolute"]

        +

        df["车组影响分"]
        * weights["group_impact"]

        +

        df["系统影响分"]
        * weights["system_impact"]

        +

        df["扩散分"]
        * weights["spread"]

        +

        df["系统权重分"]
        * weights["system_weight"]

    ) / total_weight * 100

    df["动态风险分"] = (
        df["动态风险分"]
        .round(2)
    )

    return df


# ============================================================
# 层级内部 P1/P2/P3/P4
# ============================================================

def assign_layer_priorities(
    layer_df: pd.DataFrame,
    layer_name: str,
):

    df = layer_df.copy()

    if df.empty:

        return df

    # --------------------------------------------------------
    # 待确认数据不自动进入P1
    # --------------------------------------------------------

    if layer_name == "待确认":

        df["最终优先级"] = "待确认"

        df["处理建议"] = (
            "先确认车组/通道规则，再判断系统优先级"
        )

        return df

    # --------------------------------------------------------
    # 计算风险
    # --------------------------------------------------------

    df = score_within_layer(
        df
    )

    number = len(df)

    rank = df[
        "动态风险分"
    ].rank(
        method="first",
        ascending=False,
    )

    p1 = dynamic_cut(
        number,
        PRIORITY_SHARE["P1"],
    )

    p2 = max(
        p1,
        dynamic_cut(
            number,
            PRIORITY_SHARE["P2"],
        ),
    )

    p3 = max(
        p2,
        dynamic_cut(
            number,
            PRIORITY_SHARE["P3"],
        ),
    )

    # 默认P4
    df["最终优先级"] = "P4"

    df.loc[
        rank <= p1,
        "最终优先级"
    ] = "P1"

    df.loc[
        (rank > p1)
        &
        (rank <= p2),
        "最终优先级"
    ] = "P2"

    df.loc[
        (rank > p2)
        &
        (rank <= p3),
        "最终优先级"
    ] = "P3"

    df["处理建议"] = (
        df["最终优先级"]
        .map(PRIORITY_ADVICE)
    )

    return df


# ============================================================
# 最终优先级
#
# V7最核心的算法
#
# 核心系统
#     ↓
# 普通摄像头
#     ↓
# 待确认
#
# 所以普通摄像头即使异常次数非常多，
# 也不会跨层超过DMS/ADAS/DSC。
# ============================================================

def assign_final_priority(
    stats: pd.DataFrame,
):

    if stats.empty:

        empty = stats.copy()

        empty["动态风险分"] = (
            pd.Series(
                dtype=float
            )
        )

        empty["最终优先级"] = (
            pd.Series(
                dtype=str
            )
        )

        empty["处理建议"] = (
            pd.Series(
                dtype=str
            )
        )

        summary = pd.DataFrame({

            "系统层级": [
                "核心系统",
                "普通摄像头",
                "待确认",
            ],

            "P1": [0, 0, 0],

            "P2": [0, 0, 0],

            "P3": [0, 0, 0],

            "P4": [0, 0, 0],
        })

        return empty, summary

    frames = []

    # ========================================================
    # 核心系统
    # ========================================================

    core = stats[
        stats["系统层级"]
        ==
        "核心系统"
    ].copy()

    frames.append(
        assign_layer_priorities(
            core,
            "核心系统",
        )
    )

    # ========================================================
    # 普通摄像头
    # ========================================================

    normal = stats[
        stats["系统层级"]
        ==
        "普通摄像头"
    ].copy()

    frames.append(
        assign_layer_priorities(
            normal,
            "普通摄像头",
        )
    )

    # ========================================================
    # 待确认
    # ========================================================

    review = stats[
        stats["系统层级"]
        ==
        "待确认"
    ].copy()

    frames.append(
        assign_layer_priorities(
            review,
            "待确认",
        )
    )

    result = pd.concat(
        frames,
        ignore_index=True,
    )

    # ========================================================
    # 系统层级排序
    # ========================================================

    layer_order = {

        "核心系统": 1,

        "普通摄像头": 2,

        "待确认": 3,
    }

    result["_layer_order"] = (

        result["系统层级"]

        .map(layer_order)

        .fillna(9)
    )

    result["_priority_order"] = (

        result["最终优先级"]

        .map(PRIORITY_ORDER)

        .fillna(99)
    )

    # ========================================================
    # 最终排序
    #
    # 第一层：核心系统
    # 第二层：普通摄像头
    # 第三层：待确认
    #
    # 同层级内再看P1/P2/P3/P4
    # ========================================================

    result = (

        result

        .sort_values(

            [
                "_layer_order",

                "_priority_order",

                "动态风险分",

                "视频丢失次数",

                "归属车组",

                "通道号",
            ],

            ascending=[

                True,

                True,

                False,

                False,

                True,

                True,
            ],

            na_position="last",
        )

        .drop(
            columns=[
                "_layer_order",
                "_priority_order",
            ]
        )
    )

    # ========================================================
    # 层级优先级统计
    # ========================================================

    rows = []

    for layer in [

        "核心系统",

        "普通摄像头",

        "待确认",
    ]:

        part = result[
            result["系统层级"]
            ==
            layer
        ]

        rows.append({

            "系统层级": layer,

            "P1": int(
                (
                    part["最终优先级"]
                    ==
                    "P1"
                ).sum()
            ),

            "P2": int(
                (
                    part["最终优先级"]
                    ==
                    "P2"
                ).sum()
            ),

            "P3": int(
                (
                    part["最终优先级"]
                    ==
                    "P3"
                ).sum()
            ),

            "P4": int(
                (
                    part["最终优先级"]
                    ==
                    "P4"
                ).sum()
            ),
        })

    summary = pd.DataFrame(
        rows
    )

    return result, summary


# ============================================================
# 系统统计
# ============================================================

def make_system_summary(
    stats: pd.DataFrame,
):

    if stats.empty:

        return pd.DataFrame(
            columns=[
                "系统层级",
                "系统类型",
                "视频丢失次数",
            ]
        )

    return (

        stats

        .groupby(
            [
                "系统层级",
                "系统类型",
            ],
            dropna=False,
        )["视频丢失次数"]

        .sum()

        .reset_index()

        .sort_values(
            [
                "系统层级",
                "视频丢失次数",
            ],

            ascending=[
                True,
                False,
            ],
        )
    )


# ============================================================
# 车组统计
# ============================================================

def make_group_summary(
    stats: pd.DataFrame,
):

    if stats.empty:

        return pd.DataFrame(
            columns=[
                "归属车组",
                "系统层级",
                "视频丢失次数",
            ]
        )

    return (

        stats

        .groupby(
            [
                "归属车组",
                "系统层级",
            ],
            dropna=False,
        )["视频丢失次数"]

        .sum()

        .reset_index()

        .sort_values(
            "视频丢失次数",
            ascending=False,
        )
    )


# ============================================================
# 写Excel
# ============================================================

def write_output(
    raw_count: int,
    detail: pd.DataFrame,
    stats: pd.DataFrame,
    priority: pd.DataFrame,
    priority_summary: pd.DataFrame,
):

    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    temp = (
        OUTPUT_DIR
        /
        TEMP_FILE_NAME
    )

    output = (
        OUTPUT_DIR
        /
        OUTPUT_FILE_NAME
    )

    if temp.exists():

        try:

            temp.unlink()

        except OSError:

            pass

    # ========================================================
    # 1. 总体系统统计
    # ========================================================

    system_summary = (
        make_system_summary(
            stats
        )
    )

    # ========================================================
    # 2. 车组异常统计
    # ========================================================

    group_summary = (
        make_group_summary(
            stats
        )
    )

    # ========================================================
    # 3. DMS / ADAS / DSC
    # ========================================================

    core = stats[
        stats["系统层级"]
        ==
        "核心系统"
    ].copy()

    if core.empty:

        core_summary = pd.DataFrame(
            columns=[
                "系统类型",
                "归属车组",
                "视频丢失次数",
            ]
        )

    else:

        core_summary = (

            core

            .groupby(
                [
                    "系统类型",
                    "归属车组",
                ],

                dropna=False,
            )["视频丢失次数"]

            .sum()

            .reset_index()

            .sort_values(
                "视频丢失次数",
                ascending=False,
            )
        )

    # ========================================================
    # 4. 最高优先级重点
    # ========================================================

    focus = (

        priority[
            priority["最终优先级"]
            .isin(
                [
                    "P1",
                    "P2",
                    "P3",
                    "P4",
                ]
            )
        ]

        .head(
            TOP_FOCUS_COUNT
        )

        .copy()
    )

    # ========================================================
    # 5. TOP100
    # ========================================================

    top100 = (
        priority
        .head(TOP100_COUNT)
        .copy()
    )

    # ========================================================
    # 6. 待确认
    # ========================================================

    review = priority[
        priority["最终优先级"]
        ==
        "待确认"
    ].copy()

    # ========================================================
    # 7. 数据质量
    # ========================================================

    quality = pd.DataFrame({

        "检查项目": [

            "原始记录数",

            "视频/摄像头异常记录数",

            "异常车组-通道组合数",

            "核心系统异常车组-通道数",

            "普通摄像头异常车组-通道数",

            "待确认车组-通道数",
        ],

        "结果": [

            raw_count,

            len(detail),

            len(stats),

            int(
                (
                    stats["系统层级"]
                    ==
                    "核心系统"
                ).sum()
            )
            if len(stats)
            else 0,

            int(
                (
                    stats["系统层级"]
                    ==
                    "普通摄像头"
                ).sum()
            )
            if len(stats)
            else 0,

            int(
                (
                    stats["系统层级"]
                    ==
                    "待确认"
                ).sum()
            )
            if len(stats)
            else 0,
        ],
    })

    # ========================================================
    # 8. 规则覆盖检查
    # ========================================================

    if not stats.empty:

        rule_check = (

            stats

            .groupby(
                [
                    "系统层级",
                    "系统类型",
                    "分类依据",
                    "规则状态",
                ],

                dropna=False,
            )

            .agg(

                视频丢失次数=(
                    "视频丢失次数",
                    "sum",
                ),

                车组数=(
                    "归属车组",
                    "nunique",
                ),

                通道数=(
                    "通道号",
                    "nunique",
                ),
            )

            .reset_index()

            .sort_values(
                "视频丢失次数",
                ascending=False,
            )
        )

    else:

        rule_check = pd.DataFrame(
            columns=[
                "系统层级",
                "系统类型",
                "分类依据",
                "规则状态",
                "视频丢失次数",
                "车组数",
                "通道数",
            ]
        )

    # ========================================================
    # 9. V7规则说明
    # ========================================================

    priority_explanation = pd.DataFrame({

        "规则": [

            "第一层：系统优先级",

            "核心系统",

            "普通摄像头",

            "待确认",

            "第二层：同层级风险排序",

            "异常次数",

            "车组总体影响",

            "系统总体影响",

            "异常通道扩散",
        ],

        "V7规则": [

            "系统层级优先于异常次数",

            "DMS / ADAS / DSC",

            "盲区摄像头 / 前摄像机 / 原车倒车 / 未接摄像机",

            "无法可靠映射的车组-通道",

            "只在同一层级内部比较",

            "越多风险越高",

            "同一车组异常越集中风险越高",

            "同一系统整体异常越多风险越高",

            "同一车组异常通道越多风险越高",
        ],
    })

    # ========================================================
    # 写Excel
    # ========================================================

    with pd.ExcelWriter(
        temp,
        engine="openpyxl",
        mode="w",
    ) as writer:

        sheets = [

            (
                "总体系统统计",
                system_summary,
            ),

            (
                "车组异常统计",
                group_summary,
            ),

            (
                "DMS_ADAS_DSC统计",
                core_summary,
            ),

            (
                "最高优先级重点",
                focus,
            ),

            (
                "动态优先级TOP100",
                top100,
            ),

            (
                "完整动态优先级",
                priority,
            ),

            (
                "优先级层级统计",
                priority_summary,
            ),

            (
                "待确认规则",
                review,
            ),

            (
                "规则覆盖检查",
                rule_check,
            ),

            (
                "摄像头异常明细",
                detail,
            ),

            (
                "数据质量检查",
                quality,
            ),

            (
                "优先级规则说明",
                priority_explanation,
            ),
        ]

        for sheet_name, frame in sheets:

            frame.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
            )

        # ====================================================
        # Excel格式处理
        # ====================================================

        from openpyxl.styles import Font

        workbook = writer.book

        if not workbook.worksheets:

            workbook.create_sheet(
                "数据质量检查"
            )

        for worksheet in workbook.worksheets:

            # 强制可见
            worksheet.sheet_state = "visible"

            # 冻结首行
            worksheet.freeze_panes = "A2"

            # 自动筛选
            if (
                worksheet.max_row >= 1
                and worksheet.max_column >= 1
            ):

                worksheet.auto_filter.ref = (
                    worksheet.dimensions
                )

            # 首行加粗
            for cell in worksheet[1]:

                cell.font = Font(
                    bold=True
                )

            # 自动列宽
            for column_cells in worksheet.iter_cols():

                values = [

                    ""

                    if cell.value is None

                    else str(cell.value)

                    for cell
                    in list(column_cells)[:200]
                ]

                max_length = max(
                    (
                        len(value)
                        for value in values
                    ),
                    default=8,
                )

                width = min(
                    max(
                        max_length + 2,
                        10,
                    ),
                    42,
                )

                worksheet.column_dimensions[
                    column_cells[0].column_letter
                ].width = width

    # ========================================================
    # 再次验证Excel
    # ========================================================

    from openpyxl import load_workbook

    workbook = load_workbook(
        temp
    )

    if not workbook.worksheets:

        workbook.create_sheet(
            "数据质量检查"
        )

    visible_exists = any(

        worksheet.sheet_state
        ==
        "visible"

        for worksheet
        in workbook.worksheets
    )

    if not visible_exists:

        workbook.worksheets[0].sheet_state = (
            "visible"
        )

    workbook.save(
        temp
    )

    workbook.close()

    # ========================================================
    # 替换正式文件
    # ========================================================

    try:

        os.replace(
            temp,
            output,
        )

        return output

    except PermissionError:

        stamp = time.strftime(
            "%Y%m%d_%H%M%S"
        )

        fallback = (

            OUTPUT_DIR
            /
            f"视频异常巡检最终汇总_V7_{stamp}.xlsx"
        )

        os.replace(
            temp,
            fallback,
        )

        return fallback


# ============================================================
# 主程序
# ============================================================

def main():

    print()
    print("=" * 78)

    print(
        "视频异常巡检 V7"
    )

    print(
        "DMS / ADAS / DSC 核心系统优先版"
    )

    print("=" * 78)

    # ========================================================
    # 读取数据
    # ========================================================

    raw, input_path = (
        load_data()
    )

    print(
        f"输入文件：{input_path}"
    )

    print(
        f"原始记录：{len(raw)}"
    )

    # ========================================================
    # 提取视频异常
    # ========================================================

    detail = (
        extract_camera_errors(
            raw
        )
    )

    # ========================================================
    # 统计
    # ========================================================

    stats = (
        build_stats(
            detail
        )
    )

    # ========================================================
    # 优先级
    # ========================================================

    priority, priority_summary = (
        assign_final_priority(
            stats
        )
    )

    print()
    print(
        f"视频/摄像头异常记录：{len(detail)}"
    )

    print(
        f"异常车组-通道组合：{len(stats)}"
    )

    # ========================================================
    # 系统统计
    # ========================================================

    print()
    print(
        "--- 系统层级 ---"
    )

    if len(stats):

        system_layer_summary = (

            stats

            .groupby(
                [
                    "系统层级",
                    "系统类型",
                ]
            )["视频丢失次数"]

            .sum()

            .sort_values(
                ascending=False
            )
        )

        print(
            system_layer_summary.to_string()
        )

    else:

        print(
            "没有识别到视频异常。"
        )

    # ========================================================
    # 优先级统计
    # ========================================================

    print()
    print(
        "--- 优先级层级统计 ---"
    )

    print(
        priority_summary.to_string(
            index=False
        )
    )

    # ========================================================
    # 最高优先级
    # ========================================================

    print()
    print(
        "--- V7 最高优先级重点（前20）---"
    )

    if len(priority):

        columns = [

            "归属车组",

            "通道号",

            "系统层级",

            "系统类型",

            "视频丢失次数",

            "动态风险分",

            "最终优先级",

            "处理建议",
        ]

        print(

            priority[
                columns
            ]

            .head(20)

            .to_string(
                index=False
            )
        )

    # ========================================================
    # 输出Excel
    # ========================================================

    output = write_output(

        len(raw),

        detail,

        stats,

        priority,

        priority_summary,
    )

    print()
    print(
        f"结果文件：{output}"
    )

    print()
    print(
        "程序执行完成。"
    )

    print("=" * 78)


# ============================================================
# 程序入口
# ============================================================

if __name__ == "__main__":

    try:

        main()

    except Exception as exc:

        print()
        print(
            f"程序运行失败："
            f"{type(exc).__name__}: {exc}"
        )

        raise